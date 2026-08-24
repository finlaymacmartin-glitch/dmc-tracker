"""
Delisle Mowing Co. -- import a DMC app export (JSON) into DMC_Books.xlsx.

Usage:
    python import_to_excel.py <export.json> [--workbook PATH] [--replace]

Behavior:
  * Creates the workbook if it doesn't exist.
  * Raw data sheets (Clients, Contracts, Invoices, Payments, Expenses, Budgets)
    are merged by record UUID -- the incoming record wins when its updatedAt is
    newer. Safe to import the same file twice (no duplicates).
  * --replace makes the raw sheets an exact snapshot of the export instead
    (records deleted in the app disappear from the workbook too).
  * Computed sheets (AR Aging, Budget vs Actual, Summary) are rebuilt from the
    merged raw data on every run.

Requires: openpyxl  (pip install openpyxl)
"""

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

GREEN = "1B5E20"
LIGHT_GREEN = "E8F5E9"
RED = "C62828"
CAD = '"$"#,##0.00'

# ---------------------------------------------------------------- sheet specs
# (sheet name, entity key in export, [(header, record field)])
RAW_SHEETS = [
    ("Clients", "clients", [
        ("ID", "id"), ("Name", "name"), ("Phone", "phone"), ("Email", "email"),
        ("Address", "address"), ("Notes", "notes"), ("Updated At", "updatedAt"),
    ]),
    ("Contracts", "contracts", [
        ("ID", "id"), ("Client ID", "clientId"), ("Service", "service"),
        ("Description", "description"), ("Price", "price"), ("Billing", "billing"),
        ("Frequency", "frequency"), ("Start Date", "startDate"), ("End Date", "endDate"),
        ("Status", "status"), ("Notes", "notes"), ("Updated At", "updatedAt"),
    ]),
    ("Invoices", "invoices", [
        ("ID", "id"), ("Invoice #", "number"), ("Client ID", "clientId"),
        ("Contract ID", "contractId"), ("Date Issued", "dateIssued"), ("Due Date", "dueDate"),
        ("Amount", "amount"), ("Status", "status"), ("Notes", "notes"), ("Updated At", "updatedAt"),
    ]),
    ("Payments", "payments", [
        ("ID", "id"), ("Invoice ID", "invoiceId"), ("Client ID", "clientId"),
        ("Date", "date"), ("Amount", "amount"), ("Method", "method"),
        ("Note", "note"), ("Updated At", "updatedAt"),
    ]),
    ("Visits", "visits", [
        ("ID", "id"), ("Contract ID", "contractId"), ("Client ID", "clientId"),
        ("Date", "date"), ("Note", "note"), ("Invoice ID", "invoiceId"), ("Updated At", "updatedAt"),
    ]),
    ("Expenses", "expenses", [
        ("ID", "id"), ("Date", "date"), ("Amount", "amount"), ("Category", "category"),
        ("Vendor", "vendor"), ("Line", "line"), ("Note", "note"), ("Updated At", "updatedAt"),
    ]),
    ("Budgets", "budgets", [
        ("ID", "id"), ("Category", "category"), ("Period", "period"),
        ("Limit", "limit"), ("Updated At", "updatedAt"),
    ]),
    ("Quotes", "quotes", [
        ("ID", "id"), ("Client ID", "clientId"), ("Prospect", "prospectName"),
        ("Service", "service"), ("Description", "description"), ("Price", "price"),
        ("Billing", "billing"), ("Date Sent", "dateIssued"), ("Expires", "expiryDate"),
        ("Status", "status"), ("Updated At", "updatedAt"),
    ]),
    ("Mileage", "mileage", [
        ("ID", "id"), ("Date", "date"), ("KM", "km"), ("Line", "line"),
        ("Note", "note"), ("Updated At", "updatedAt"),
    ]),
    ("Equipment", "equipment", [
        ("ID", "id"), ("Name", "name"), ("Purchase Date", "purchaseDate"), ("Cost", "cost"),
        ("Line", "line"), ("Last Service", "lastServiceDate"), ("Service Notes", "serviceNotes"),
        ("Notes", "notes"), ("Updated At", "updatedAt"),
    ]),
    ("Crew", "crew", [
        ("ID", "id"), ("Name", "name"), ("Phone", "phone"), ("Default Rate", "defaultRate"),
        ("Status", "status"), ("Notes", "notes"), ("Updated At", "updatedAt"),
    ]),
    ("Shifts", "shifts", [
        ("ID", "id"), ("Crew ID", "crewId"), ("Date", "date"), ("Hours", "hours"),
        ("Rate", "rate"), ("Flat", "flatAmount"), ("Amount", "amount"), ("Client ID", "clientId"),
        ("Line", "line"), ("Paid", "paid"), ("Paid Date", "paidDate"), ("Note", "note"),
        ("Updated At", "updatedAt"),
    ]),
]
CURRENCY_HEADERS = {"Amount", "Price", "Limit", "Cost", "Default Rate", "Rate", "Flat"}
NUMERIC_HEADERS = {"KM", "Hours"}
COMPUTED_SHEETS = ["AR Aging", "Budget vs Actual", "Labour", "Summary"]


def is_paid(shift):
    return str(shift.get("paid", "")).strip().lower() == "true"


# ---------------------------------------------------------------- helpers
def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def read_existing(ws, columns):
    """Read a raw sheet back into {id: record} using its header row."""
    records = {}
    if ws.max_row < 2:
        return records
    headers = [cell_str(c.value) for c in ws[1]]
    header_to_field = {h: f for h, f in columns}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for header, value in zip(headers, row):
            f = header_to_field.get(header)
            if not f:
                continue
            if header in CURRENCY_HEADERS or header in NUMERIC_HEADERS:
                rec[f] = float(value or 0)
            else:
                rec[f] = cell_str(value)
        if rec.get("id"):
            records[rec["id"]] = rec
    return records


def merge(existing, incoming):
    """Incoming record wins when its updatedAt is newer (or existing lacks one)."""
    merged = dict(existing)
    for rid, rec in incoming.items():
        old = merged.get(rid)
        if old is None or cell_str(rec.get("updatedAt")) >= cell_str(old.get("updatedAt")):
            merged[rid] = rec
    return merged


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_raw_sheet(wb, name, columns, records):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append([h for h, _ in columns])
    style_header(ws, len(columns))
    rows = sorted(records.values(), key=lambda r: cell_str(r.get("date") or r.get("dateIssued") or r.get("name") or ""))
    for rec in rows:
        ws.append([rec.get(f, "") for _, f in columns])
    for col_idx, (header, _) in enumerate(columns, start=1):
        if header in CURRENCY_HEADERS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = CAD
    autosize(ws, [38 if h.endswith("ID") or h == "ID" else 16 for h, _ in columns])
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(2, ws.max_row)}"
    return ws


# ---------------------------------------------------------------- derived math
def invoice_state(inv, payments):
    paid = round(sum(float(p.get("amount") or 0) for p in payments if p.get("invoiceId") == inv["id"]), 2)
    amount = float(inv.get("amount") or 0)
    balance = max(0.0, round(amount - paid, 2))
    days_past_due = 0
    if inv.get("dueDate") and balance > 0.004:
        try:
            days_past_due = (date.today() - datetime.strptime(inv["dueDate"], "%Y-%m-%d").date()).days
        except ValueError:
            days_past_due = 0
    if inv.get("status") == "draft":
        status = "draft"
    elif balance <= 0.004 and amount > 0:
        status = "paid"
    elif days_past_due > 0:
        status = "overdue"
    elif paid > 0:
        status = "partial"
    else:
        status = "sent"
    return paid, balance, max(0, days_past_due), status


def month_key(date_str):
    return (date_str or "")[:7]


def season_months(ref: date):
    """Mowing season Apr-Oct, plowing season Nov-Mar (spans year end)."""
    if 4 <= ref.month <= 10:
        return [f"{ref.year}-{m:02d}" for m in range(4, 11)]
    start_year = ref.year if ref.month >= 11 else ref.year - 1
    return [f"{start_year}-11", f"{start_year}-12",
            f"{start_year + 1}-01", f"{start_year + 1}-02", f"{start_year + 1}-03"]


def quote_winrate(quotes):
    won = sum(1 for q in quotes.values() if q.get("status") == "accepted")
    decided = sum(1 for q in quotes.values() if q.get("status") in ("accepted", "declined"))
    return f"{won} / {decided}" if decided else "n/a"


def title_row(ws, row, text, ncols=1):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12, color=GREEN)
    return row + 1


def header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=GREEN)
    return row + 1


# ---------------------------------------------------------------- computed sheets
def build_ar_aging(wb, data):
    if "AR Aging" in wb.sheetnames:
        del wb["AR Aging"]
    ws = wb.create_sheet("AR Aging")
    clients = data["clients"]
    invoices = [i for i in data["invoices"].values() if i.get("status") != "draft"]
    payments = list(data["payments"].values())

    row = title_row(ws, 1, f"AR Aging — as of {date.today().isoformat()}")
    row += 1
    row = header_row(ws, row, ["Client", "Not Yet Due", "1–30 Days", "31–60 Days", "61–90 Days", "90+ Days", "Total Owing"])
    totals = [0.0] * 6
    for cid, client in sorted(clients.items(), key=lambda kv: kv[1].get("name", "")):
        buckets = [0.0] * 5
        for inv in invoices:
            if inv.get("clientId") != cid:
                continue
            _, balance, dpd, _ = invoice_state(inv, payments)
            if balance <= 0.004:
                continue
            idx = 0 if dpd <= 0 else 1 if dpd <= 30 else 2 if dpd <= 60 else 3 if dpd <= 90 else 4
            buckets[idx] += balance
        total = round(sum(buckets), 2)
        if total <= 0.004:
            continue
        ws.append([client.get("name", "?")] + [round(b, 2) for b in buckets] + [total])
        for i, b in enumerate(buckets):
            totals[i] += b
        totals[5] += total
        row += 1
    ws.append(["TOTAL"] + [round(t, 2) for t in totals])
    for c in ws[row]:
        c.font = Font(bold=True)
    row += 2

    # open invoice detail
    row = title_row(ws, row, "Open invoices")
    row = header_row(ws, row, ["Invoice #", "Client", "Issued", "Due", "Amount", "Paid", "Balance", "Days Past Due", "Status"])
    open_rows = []
    for inv in invoices:
        paid, balance, dpd, status = invoice_state(inv, payments)
        if balance <= 0.004:
            continue
        open_rows.append((dpd, [inv.get("number", ""), clients.get(inv.get("clientId"), {}).get("name", "?"),
                                inv.get("dateIssued", ""), inv.get("dueDate", ""),
                                float(inv.get("amount") or 0), paid, balance, dpd, status]))
    for dpd, vals in sorted(open_rows, key=lambda t: -t[0]):
        ws.append(vals)
        if vals[8] == "overdue":
            for c in ws[ws.max_row]:
                c.font = Font(color=RED)

    for col_letter in ("B", "C", "D", "E", "F", "G"):
        for cell in ws[col_letter]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = CAD
    autosize(ws, [24, 14, 14, 14, 14, 14, 14, 15, 12])


def build_budget_sheet(wb, data):
    if "Budget vs Actual" in wb.sheetnames:
        del wb["Budget vs Actual"]
    ws = wb.create_sheet("Budget vs Actual")
    expenses = list(data["expenses"].values())
    today = date.today()
    this_month = today.strftime("%Y-%m")
    season = season_months(today)

    row = title_row(ws, 1, f"Budget vs Actual — {this_month}")
    row += 1
    row = header_row(ws, row, ["Category", "Period", "Limit", "Actual", "Remaining", "% Used"])
    for b in sorted(data["budgets"].values(), key=lambda x: x.get("category", "")):
        limit = float(b.get("limit") or 0)
        months = [this_month] if b.get("period") == "monthly" else season
        actual = round(sum(float(e.get("amount") or 0) for e in expenses
                           if e.get("category") == b.get("category") and month_key(e.get("date")) in months), 2)
        pct = actual / limit if limit else 0
        ws.append([b.get("category"), b.get("period"), limit, actual, round(limit - actual, 2), pct])
        r = ws.max_row
        for col in ("C", "D", "E"):
            ws[f"{col}{r}"].number_format = CAD
        ws[f"F{r}"].number_format = "0%"
        if pct >= 1:
            ws[f"F{r}"].font = Font(bold=True, color=RED)
        row += 1
    row += 2

    # expenses by category x month (last 12 months)
    months = sorted({month_key(e.get("date")) for e in expenses if e.get("date")})[-12:]
    categories = sorted({e.get("category", "Other") for e in expenses})
    row = title_row(ws, row, "Expenses by category and month")
    row = header_row(ws, row, ["Category"] + months + ["Total"])
    for cat in categories:
        vals = [round(sum(float(e.get("amount") or 0) for e in expenses
                          if e.get("category") == cat and month_key(e.get("date")) == m), 2) for m in months]
        ws.append([cat] + vals + [round(sum(vals), 2)])
        for c in ws[ws.max_row][1:]:
            c.number_format = CAD
    if categories:
        totals = [round(sum(float(e.get("amount") or 0) for e in expenses
                            if month_key(e.get("date")) == m), 2) for m in months]
        ws.append(["TOTAL"] + totals + [round(sum(totals), 2)])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
            if isinstance(c.value, (int, float)):
                c.number_format = CAD
    autosize(ws, [24] + [12] * (len(months) + 5))


def build_labour(wb, data):
    if "Labour" in wb.sheetnames:
        del wb["Labour"]
    ws = wb.create_sheet("Labour")
    crew = data.get("crew", {})
    shifts = list(data.get("shifts", {}).values())
    payments = list(data["payments"].values())
    expenses = list(data["expenses"].values())

    row = title_row(ws, 1, f"Labour analysis — as of {date.today().isoformat()}")
    row += 1

    # ---- owed by crew member ----
    row = title_row(ws, row, "Owed to crew (unpaid shifts)")
    row = header_row(ws, row, ["Crew Member", "Unpaid Shifts", "Owed"])
    total_owed = 0.0
    for cid, member in sorted(crew.items(), key=lambda kv: kv[1].get("name", "")):
        unpaid = [s for s in shifts if s.get("crewId") == cid and not is_paid(s)]
        owed = round(sum(float(s.get("amount") or 0) for s in unpaid), 2)
        if owed <= 0:
            continue
        ws.append([member.get("name", "?"), len(unpaid), owed])
        ws.cell(row=ws.max_row, column=3).number_format = CAD
        total_owed += owed
        row += 1
    ws.append(["TOTAL OWED", "", round(total_owed, 2)])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3).number_format = CAD
    row = ws.max_row + 3

    # ---- wages paid by month + labour % of that month's revenue ----
    row = title_row(ws, row, "Wages paid by month")
    row = header_row(ws, row, ["Month", "Wages Paid", "Revenue Collected", "Labour %"])
    months = sorted({month_key(s.get("paidDate")) for s in shifts if is_paid(s) and s.get("paidDate")})
    for m in months:
        paid = round(sum(float(s.get("amount") or 0) for s in shifts if is_paid(s) and month_key(s.get("paidDate")) == m), 2)
        rev = round(sum(float(p.get("amount") or 0) for p in payments if month_key(p.get("date")) == m), 2)
        ws.append([m, paid, rev, paid / rev if rev else ""])
        r = ws.max_row
        ws[f"B{r}"].number_format = CAD
        ws[f"C{r}"].number_format = CAD
        ws[f"D{r}"].number_format = "0%"
        row += 1
    row = ws.max_row + 3

    # ---- wages by service line ----
    row = title_row(ws, row, "Wages by service line (paid + owed)")
    row = header_row(ws, row, ["Line", "Paid", "Owed", "Total"])
    for line in ("mowing", "plowing", "general"):
        paid = round(sum(float(s.get("amount") or 0) for s in shifts if is_paid(s) and (s.get("line") or "general") == line), 2)
        owed = round(sum(float(s.get("amount") or 0) for s in shifts if not is_paid(s) and (s.get("line") or "general") == line), 2)
        ws.append([line.capitalize(), paid, owed, round(paid + owed, 2)])
        for col in ("B", "C", "D"):
            ws[f"{col}{ws.max_row}"].number_format = CAD
    row = ws.max_row + 3

    # ---- hiring-power snapshot (same math as the in-app calculator) ----
    rates = [float(c.get("defaultRate") or 0) for c in crew.values() if c.get("status") == "active" and float(c.get("defaultRate") or 0) > 0]
    rate = round(sum(rates) / len(rates), 2) if rates else 20.0
    hours_wk = 10
    monthly_cost = round(rate * hours_wk * 4.33, 2)
    cur = date.today().strftime("%Y-%m")
    month_nets = {}
    for p in payments:
        m = month_key(p.get("date"))
        if m and m < cur:
            month_nets.setdefault(m, 0.0)
            month_nets[m] += float(p.get("amount") or 0)
    for e in expenses:
        m = month_key(e.get("date"))
        if m and m < cur:
            month_nets.setdefault(m, 0.0)
            month_nets[m] -= float(e.get("amount") or 0)
    last3 = [month_nets[m] for m in sorted(month_nets)[-3:]]
    avg_net = round(sum(last3) / len(last3), 2) if last3 else 0.0
    visits = len(data.get("visits", {}))
    collected = sum(float(p.get("amount") or 0) for p in payments)
    avg_visit = round(collected / visits, 2) if visits else 0.0
    weekly_cost = round(rate * hours_wk, 2)

    row = title_row(ws, row, f"Hiring-power snapshot (at {rate:.2f}/hr, {hours_wk} hrs/week)")
    for label, value, fmt in [
        ("Helper cost per month", monthly_cost, CAD),
        (f"Avg net per month (last {len(last3)} completed)", avg_net, CAD),
        ("Helpers current profit covers", int(avg_net // monthly_cost) if monthly_cost and avg_net > 0 else 0, "0"),
        ("Avg revenue per visit", avg_visit, CAD),
        ("Break-even extra visits/week per helper", round(weekly_cost / avg_visit, 1) if avg_visit else "n/a", "0.0"),
    ]:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=value)
        if isinstance(value, (int, float)):
            c.number_format = fmt
        c.font = Font(bold=True)
        row += 1
    autosize(ws, [38, 18, 20, 12])


def build_summary(wb, data):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)
    payments = list(data["payments"].values())
    expenses = list(data["expenses"].values())
    invoices = data["invoices"]
    contracts = data["contracts"]

    def payment_line(p):
        """Map a payment to a service line via its invoice's contract."""
        inv = invoices.get(p.get("invoiceId"), {})
        contract = contracts.get(inv.get("contractId"), {})
        service = contract.get("service", "")
        return service if service in ("mowing", "plowing") else "general"

    row = title_row(ws, 1, f"Delisle Mowing Co. — Summary (built {date.today().isoformat()})")
    ws.cell(row=2, column=1, value="Cash basis: revenue = payments received. See AR Aging for money still owed.").font = Font(italic=True, size=9)
    row = 4

    # ---- monthly P&L ----
    row = title_row(ws, row, "Monthly cash P&L")
    months = sorted({month_key(p.get("date")) for p in payments if p.get("date")} |
                    {month_key(e.get("date")) for e in expenses if e.get("date")})
    row = header_row(ws, row, ["Month", "Revenue Collected", "Expenses", "Net"])
    for m in months:
        rev = round(sum(float(p.get("amount") or 0) for p in payments if month_key(p.get("date")) == m), 2)
        exp = round(sum(float(e.get("amount") or 0) for e in expenses if month_key(e.get("date")) == m), 2)
        ws.append([m, rev, exp, round(rev - exp, 2)])
        r = ws.max_row
        for col in ("B", "C", "D"):
            ws[f"{col}{r}"].number_format = CAD
        if rev - exp < 0:
            ws[f"D{r}"].font = Font(color=RED)
        row += 1
    total_rev = round(sum(float(p.get("amount") or 0) for p in payments), 2)
    total_exp = round(sum(float(e.get("amount") or 0) for e in expenses), 2)
    ws.append(["TOTAL", total_rev, total_exp, round(total_rev - total_exp, 2)])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
        if isinstance(c.value, (int, float)):
            c.number_format = CAD
    row = ws.max_row + 3

    # ---- by service line (the snowplow-decision view) ----
    row = title_row(ws, row, "By service line (all time)")
    row = header_row(ws, row, ["Line", "Revenue Collected", "Expenses", "Net"])
    for line in ("mowing", "plowing", "general"):
        rev = round(sum(float(p.get("amount") or 0) for p in payments if payment_line(p) == line), 2)
        exp = round(sum(float(e.get("amount") or 0) for e in expenses if (e.get("line") or "general") == line), 2)
        ws.append([line.capitalize(), rev, exp, round(rev - exp, 2)])
        r = ws.max_row
        for col in ("B", "C", "D"):
            ws[f"{col}{r}"].number_format = CAD
    row = ws.max_row + 3

    # ---- headline stats ----
    open_balance = round(sum(invoice_state(i, payments)[1] for i in invoices.values() if i.get("status") != "draft"), 2)
    row = title_row(ws, row, "Headline")
    for label, value, fmt in [
        ("Clients", len(data["clients"]), "0"),
        ("Active contracts", sum(1 for k in contracts.values() if k.get("status") == "active"), "0"),
        ("Visits logged", len(data.get("visits", {})), "0"),
        ("Mileage logged (km)", round(sum(float(m.get("km") or 0) for m in data.get("mileage", {}).values()), 1), "0.0"),
        ("Equipment invested", round(sum(float(q.get("cost") or 0) for q in data.get("equipment", {}).values()), 2), CAD),
        ("Quotes won / decided", quote_winrate(data.get("quotes", {})), "@"),
        ("Wages owed to crew", round(sum(float(s.get("amount") or 0) for s in data.get("shifts", {}).values() if not is_paid(s)), 2), CAD),
        ("Labour % of revenue (all time)", (
            round(sum(float(s.get("amount") or 0) for s in data.get("shifts", {}).values()) / total_rev, 4)
            if total_rev else "n/a"), "0.0%"),
        ("AR outstanding", open_balance, CAD),
        ("All-time revenue collected", total_rev, CAD),
        ("All-time expenses", total_exp, CAD),
        ("All-time net (cash)", round(total_rev - total_exp, 2), CAD),
    ]:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = fmt
        c.font = Font(bold=True)
        row += 1
    autosize(ws, [30, 20, 16, 16])


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Import a DMC app export into the Excel workbook.")
    ap.add_argument("export_file", help="Path to the dmc-export-*.json file from the app")
    ap.add_argument("--workbook", default="DMC_Books.xlsx", help="Workbook to create/update (default: DMC_Books.xlsx)")
    ap.add_argument("--replace", action="store_true", help="Exact snapshot: raw sheets become the export contents (deletions propagate)")
    args = ap.parse_args()

    export_path = Path(args.export_file)
    if not export_path.exists():
        sys.exit(f"Export file not found: {export_path}")
    payload = json.loads(export_path.read_text(encoding="utf-8"))
    if payload.get("app") != "dmc-export" or "data" not in payload:
        sys.exit("That file is not a DMC export (missing app/data fields).")
    incoming_data = payload["data"]

    wb_path = Path(args.workbook)
    if wb_path.exists():
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    merged = {}
    for sheet_name, entity, columns in RAW_SHEETS:
        incoming = {r["id"]: r for r in incoming_data.get(entity, []) if r.get("id")}
        if args.replace or sheet_name not in wb.sheetnames:
            records = incoming
        else:
            records = merge(read_existing(wb[sheet_name], columns), incoming)
        merged[entity] = records
        write_raw_sheet(wb, sheet_name, columns, records)
        print(f"  {sheet_name:<10} {len(incoming):>4} in export -> {len(records):>4} in workbook")

    build_ar_aging(wb, merged)
    build_budget_sheet(wb, merged)
    build_labour(wb, merged)
    build_summary(wb, merged)

    # sheet order: Summary, AR Aging, Budget vs Actual, Labour, then raw data
    order = ["Summary", "AR Aging", "Budget vs Actual", "Labour"] + [s for s, _, _ in RAW_SHEETS]
    wb._sheets.sort(key=lambda ws: order.index(ws.title) if ws.title in order else 99)

    try:
        wb.save(wb_path)
    except PermissionError:
        sys.exit(f"Cannot save {wb_path} — close it in Excel and run again.")
    print(f"\nSaved {wb_path.resolve()}")
    print(f"Export was created {payload.get('exportedAt', '?')}")


if __name__ == "__main__":
    main()
